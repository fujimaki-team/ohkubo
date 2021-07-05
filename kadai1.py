import openpyxl as excel
import datetime
import os
import json
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


def set_sheet_name(jsn, sheet1):
    '''シート名を設定する'''
    key_name = "sheet_name"
    try:
        sheet1.title = jsn[key_name]
    except KeyError:
        print("sheet_name_Error")
    pass


def set_column(cols, sheet1):
    '''列の幅を設定する'''
    key_cols = "column"
    try:
        sheet1.column_dimensions[cols[key_cols]].width = cols["width"]
    except KeyError:
        print("cols_Error")
    pass


def set_line(rows, sheet1):
    '''行の幅を設定する'''
    key_rows = "line"
    try:
        sheet1.row_dimensions[rows[key_rows]].height = rows["height"]
    except KeyError:
        print("rows_Error")
    pass


def set_merges(sheet1, range):
    '''セルの結合箇所を設定する'''
    try:
        sheet1.merge_cells(range)
    except KeyError:
        print("merge_Error")
    pass


def set_borders(borders, sheet1):
    '''枠線の場所を設定する'''
    key_border = "cells"
    border = borders["border"]
    top = border["top"]
    bottom = border["bottom"]
    left = border["left"]
    right = border["right"]
    t = Side(style=top["style"], color=top["color"])
    b = Side(style=bottom["style"], color=bottom["color"])
    le = Side(style=left["style"], color=left["color"])
    r = Side(style=right["style"], color=right["color"])
    try:
        for rows in sheet1[borders[key_border]]:
            for cell in rows:
                cell.border = Border(
                    top=t, bottom=b, left=le, right=r)
    except KeyError:
        print("border_Error")
    pass


def set_sysdate(sysdate, sheet1, key):
    '''日付を設定する'''
    if "todey" == sysdate["value"]:
        try:
            sheet1[sysdate[key]] = datetime.date.today()
        except KeyError:
            print("sysdate_Error")
    else:
        print("JSONファイルの日付書式が間違えています")
    pass


def set_title(cell, value):
    '''出力文字を設定する'''
    try:
        cell.value = value
    except KeyError:
        print("Title_Error")
    pass


def set_font(cell, name, bold, size):
    ''' フォントを設定する '''
    try:
        cell.font = Font(name=name, bold=bold, size=size)
    except KeyError:
        print("Title_Font_Error")
    pass


def set_alignment(cell, horizontal, vertical):
    ''' アラインメントを設定する '''
    try:
        cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)
    except KeyError:
        print("Title_alignment_Error")
    pass


def set_fill(cell, patternType, fgColor):
    ''' セル内の色を設定する '''
    try:
        cell.fill = PatternFill(patternType=patternType, fgColor=fgColor)
    except KeyError:
        print("Title_fill_Error")
    pass


def main(sheet1):

    with open(os.path.join('.', '表紙test.json'), encoding='utf-8') as f:
        jsn = json.load(f)

    # シート名設定
    set_sheet_name(jsn, sheet1)

    # セルの幅
    for cols in jsn["cols"]:
        set_column(cols, sheet1)

    # セルの高さ
    for rows in jsn["rows"]:
        set_line(rows, sheet1)

    # セルの結合設定
    for merge in jsn["merges"]:
        range = merge['merge']
        set_merges(sheet1, range)

    # ----枠線の場所設定
    for borders in jsn["borders"]:
        set_borders(borders, sheet1)

    # 文字表示設定
    # 日付
    key = 'coordinate'
    set_sysdate(jsn["sysdate"], sheet1, key)

    for cell1 in jsn["cells_title"]:

        cell = sheet1[cell1[key]]
        value = cell1["value"]

        # 出力文字設定
        set_title(cell, value)

        # フォント設定
        font = cell1["font"]
        set_font(cell, **font)

        # 文字上下左右そろえ
        alignment = cell1["alignment"]
        set_alignment(cell, **alignment)

        # セル内の色を設定
        fill = cell1["fill"]
        set_fill(cell, **fill)


if __name__ == "__main__":

    # ブック作成
    book = excel.Workbook()
    sheet = book.active

    # メイン処理
    main(sheet)

    # 保存
    book.save(os.path.join('.', 'test.xlsx'))
